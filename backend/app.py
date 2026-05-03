# type: ignore
from flask import Flask, request, jsonify, render_template, session, send_file, Response
from flask_cors import CORS
from datetime import datetime, date, timezone, timedelta
import qrcode
import io
import base64
import uuid
import os
import hmac
import hashlib
from functools import wraps
from database import get_db_connection
from face_utils import face_ai
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import bcrypt
from dotenv import load_dotenv
import stripe
STRIPE_SECRET_KEY = os.getenv("STRIPE_SECRET_KEY")
stripe.api_key = STRIPE_SECRET_KEY

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_ = load_dotenv(os.path.join(BASE_DIR, ".env"))

def require_env(var_name: str) -> str:
    value = os.getenv(var_name)
    if value is None or value.strip() == "":
        raise RuntimeError(f"Missing required environment variable: {var_name}")
    return value

# Email Configuration (must be provided through environment variables)
GMAIL_USER = require_env("GMAIL_USER")
GMAIL_PASSWORD = require_env("GMAIL_PASSWORD")

def send_email(to_email: str, subject: str, body: str) -> bool:
    try:
        msg = MIMEMultipart()
        msg['From'] = GMAIL_USER
        msg['To'] = to_email
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'html'))
        
        server = smtplib.SMTP('smtp.gmail.com', 587)
        _ = server.starttls()
        _ = server.login(GMAIL_USER, GMAIL_PASSWORD)
        _ = server.send_message(msg)  # type: ignore
        _ = server.quit()
        return True
    except Exception as e:
        print(f"Email error: {e}")
        return False

# Flask app setup
app = Flask(__name__, 
            template_folder=os.path.join(BASE_DIR, 'frontend', 'templates'),
            static_folder=os.path.join(BASE_DIR, 'frontend', 'static'),
            static_url_path='/static')

# PWA Support
@app.route('/manifest.json')
def serve_manifest():
    return send_file(os.path.join(BASE_DIR, 'frontend', 'static', 'manifest.json'))

@app.route('/sw.js')
def serve_sw():
    response = send_file(os.path.join(BASE_DIR, 'frontend', 'static', 'js', 'sw.js'))
    response.headers['Service-Worker-Allowed'] = '/'
    return response

# CORS setup - Allow credentials
allowed_origins = os.getenv("CORS_ORIGINS", "http://127.0.0.1:5000,http://localhost:5000")
_ = CORS(
    app,
    supports_credentials=True,
    origins=[origin.strip() for origin in allowed_origins.split(",") if origin.strip()]
)

@app.after_request
def add_no_cache_headers(response: Response) -> Response:
    # Prevent stale dashboard HTML/JS from being served by browser cache.
    if 'text/html' in response.content_type:
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response

# Required for session
app.secret_key = require_env("FLASK_SECRET_KEY")
app.config['SESSION_PERMANENT'] = True
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=8)
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
QR_SIGNING_SECRET = require_env("FLASK_SECRET_KEY")
QR_EXPIRY_SECONDS = 15

# ============================================
# ROLE PROTECTION DECORATORS
# ============================================

def role_required(allowed_roles: list[str]):  # type: ignore
    def decorator(f):  # type: ignore
        @wraps(f)  # type: ignore
        def decorated_function(*args, **kwargs):  # type: ignore[no-untyped-def]
            # Check login session
            if 'user_id' not in session:
                return jsonify({"status": "error", "message": "Login required"}), 401
            # Check role
            if session.get('role') not in allowed_roles:
                return jsonify({"status": "error", "message": "Access denied"}), 403
            return f(*args, **kwargs)  # type: ignore
        return decorated_function
    return decorator

# ============================================
# PAGE ROLE PROTECTION DECORATORS
# ============================================

def page_role_required(allowed_roles: list):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                # JSON nahi, seedha login page pe redirect
                from flask import redirect, url_for
                return redirect(url_for('index'))
            if session.get('role') not in allowed_roles:
                from flask import redirect, url_for
                return redirect(url_for('index'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# ============================================
# SERVE HTML PAGES WITH ROLE CHECK
# ============================================

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/admin_dashboard')
@page_role_required(['Admin'])
def admin_dashboard():
    return render_template('admin_dashboard.html')
 
@app.route('/teacher_dashboard')
@page_role_required(['Teacher'])
def teacher_dashboard():
    return render_template('teacher_dashboard.html')
 
@app.route('/student_dashboard')
@page_role_required(['Student'])
def student_dashboard():
    return render_template('student_dashboard.html')

# ============================================
# GET CURRENT USER
# ============================================

@app.route('/current_user', methods=['GET'])
@role_required(['Admin', 'Teacher', 'Student'])
def current_user():
    user_id = session.get('user_id')
    
    conn = get_db_connection()
    if conn is None:  # type: ignore
        return jsonify({"status": "error", "message": "Database connection failed"}), 500
    
    cursor = conn.cursor(dictionary=True)  # type: ignore
    cursor.execute(  # type: ignore
        "SELECT user_id, full_name, email, role FROM users WHERE user_id = %s",
        (user_id,)
    )
    user = cursor.fetchone()  # type: ignore
    _ = cursor.close()  # type: ignore
    conn.close()  # type: ignore
    
    if not user:
        return jsonify({"status": "error", "message": "User not found"}), 404
    
    return jsonify({
        "status": "success",
        "user": {
            "id": user["user_id"],  # type: ignore
            "name": user["full_name"],  # type: ignore
            "email": user["email"],  # type: ignore
            "role": user["role"]  # type: ignore
        }
    })

# ============================================
# LOGIN API (WITH HASHED PASSWORD)
# ============================================

@app.route('/login', methods=['POST'])
def login():
    data = request.json  # type: ignore
    email = data.get('email') if data else None  # type: ignore
    password = data.get('password') if data else None  # type: ignore
    
    if not email or not password:
        return jsonify({"status": "error", "message": "Email and password required"}), 400

    conn = get_db_connection()
    if conn is None:  # type: ignore
        return jsonify({"status": "error", "message": "Database connection failed"}), 500

    cursor = conn.cursor(dictionary=True)  # type: ignore
    cursor.execute(  # type: ignore
        "SELECT user_id, full_name, email, role, password FROM users WHERE email = %s",
        (email,)
    )
    user = cursor.fetchone()  # type: ignore
    _ = cursor.close()  # type: ignore
    conn.close()  # type: ignore
    
    if not user:
        return jsonify({"status": "error", "message": "Invalid email or password"}), 401
    
    # Verify hashed password
    try:
        if bcrypt.checkpw(password.encode('utf-8'), user["password"].encode('utf-8')):  # type: ignore
            session['user_id'] = user["user_id"]  # type: ignore
            session['role'] = user["role"]  # type: ignore
            
            return jsonify({
                "status": "success",
                "user": {
                    "id": user["user_id"],  # type: ignore
                    "name": user["full_name"],  # type: ignore
                    "email": user["email"],  # type: ignore
                    "role": user["role"]  # type: ignore
                }
            })
        else:
            return jsonify({"status": "error", "message": "Invalid email or password"}), 401
    except Exception as e:
        # If password is not hashed (plain text), compare directly
        if user['password'] == password:
            session['user_id'] = user['user_id']
            session['role'] = user['role']
            
            # Optionally update to hashed password
            hashed = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())
            conn = get_db_connection()
            cur = conn.cursor()
            cur.execute("UPDATE users SET password = %s WHERE user_id = %s", (hashed.decode('utf-8'), user['user_id']))
            conn.commit()
            cur.close()
            conn.close()
            
            return jsonify({
                "status": "success",
                "user": {
                    "id": user['user_id'],
                    "name": user['full_name'],
                    "email": user['email'],
                    "role": user['role']
                }
            })
        else:
            return jsonify({"status": "error", "message": "Invalid email or password"}), 401

# ============================================
# GET LOW ATTENDANCE STUDENTS (ADMIN)
# ============================================
@app.route('/admin_low_attendance', methods=['GET'])
@role_required(['Admin'])
def get_low_attendance():
    conn = get_db_connection()
    if conn is None:
        return jsonify({"status": "error", "message": "DB connection failed"}), 500
    
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT 
            u.user_id,
            u.full_name,
            u.email,
            c.course_code,
            c.course_name,
            COUNT(DISTINCT CASE WHEN UPPER(ar.status) = 'PRESENT' THEN ar.record_id END) as present_days,
            total_sess.total_sessions,
            ROUND(
                COUNT(DISTINCT CASE WHEN UPPER(ar.status) = 'PRESENT' THEN ar.record_id END) * 100.0 /
                NULLIF(total_sess.total_sessions, 0), 2
            ) as percentage,
            (SELECT f.status FROM fines f 
             WHERE f.student_id = u.user_id 
             AND f.course_code = c.course_code 
             ORDER BY f.issued_date DESC LIMIT 1) as fine_status
        FROM users u
        JOIN student_enrollment se ON u.user_id = se.student_id
        JOIN sections sec ON se.section_id = sec.section_id
        JOIN course_semester cs ON sec.cs_id = cs.cs_id
        JOIN courses c ON cs.course_id = c.course_id
        JOIN (
            SELECT section_id, COUNT(*) as total_sessions 
            FROM attendance_sessions 
            WHERE is_active = 0
            GROUP BY section_id
        ) total_sess ON total_sess.section_id = sec.section_id
        LEFT JOIN attendance_records ar ON ar.student_id = u.user_id AND ar.section_id = sec.section_id
        WHERE u.role = 'Student'
        GROUP BY u.user_id, u.full_name, u.email, c.course_code, c.course_name, total_sess.total_sessions
        HAVING percentage < 75
        AND NOT EXISTS (
            SELECT 1 FROM fines f
            WHERE f.student_id = u.user_id
            AND f.course_code = c.course_code
            AND f.status IN ('Pending', 'Paid')
        )
        ORDER BY percentage ASC
    """)
    students = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify({"status": "success", "students": students})

# ============================================
# ISSUE FINE (ADMIN)
# ============================================
@app.route('/admin_issue_fine', methods=['POST'])
@role_required(['Admin'])
def issue_fine():
    data = request.get_json()
    student_id = data.get('student_id')
    course_code = data.get('course_code')
    course_name = data.get('course_name')
    percentage = data.get('percentage')
    
    conn = get_db_connection()
    if conn is None:
        return jsonify({"status": "error", "message": "DB connection failed"}), 500
    
    cursor = conn.cursor(dictionary=True)
    
    # Check if fine already issued
    cursor.execute("""
        SELECT fine_id FROM fines 
        WHERE student_id = %s AND course_code = %s AND status = 'Pending'
    """, (student_id, course_code))
    existing = cursor.fetchone()
    
    if existing:
        cursor.close()
        conn.close()
        return jsonify({"status": "error", "message": "Fine already issued for this course"}), 400
    
    # Insert fine
    cursor.execute("""
        INSERT INTO fines (student_id, course_code, course_name, attendance_percentage, fine_amount, status)
        VALUES (%s, %s, %s, %s, 500.00, 'Pending')
    """, (student_id, course_code, course_name, percentage))
    conn.commit()
    
    # Get student info for email
    cursor2 = conn.cursor(dictionary=True)
    cursor2.execute("""
       SELECT u.email, u.full_name 
       FROM users u 
       WHERE u.user_id = %s
    """, (student_id,))
    student_info = cursor2.fetchone()
    cursor2.close()

    if student_info:
        email_body = f"""
        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
            <div style="background: #e74c3c; padding: 20px; text-align: center;">
               <h2 style="color: white; margin: 0;">⚠️ Attendance Fine Issued</h2>
            </div>
        <div style="padding: 30px; background: #f9f9f9;">
            <p>Dear <strong>{student_info['full_name']}</strong>,</p>
            <p>A fine has been issued due to low attendance in the following course:</p>
            <table style="width:100%; border-collapse:collapse; margin:20px 0;">
                <tr style="background:#ecf0f1;">
                    <td style="padding:10px; border:1px solid #ddd;"><strong>Course</strong></td>
                    <td style="padding:10px; border:1px solid #ddd;">{course_code} - {course_name}</td>
                </tr>
                <tr>
                    <td style="padding:10px; border:1px solid #ddd;"><strong>Attendance</strong></td>
                    <td style="padding:10px; border:1px solid #ddd; color:#e74c3c;">
                        <strong>{percentage}%</strong>
                    </td>
                </tr>
                <tr style="background:#ecf0f1;">
                    <td style="padding:10px; border:1px solid #ddd;"><strong>Fine Amount</strong></td>
                    <td style="padding:10px; border:1px solid #ddd; color:#e74c3c;">
                        <strong>Rs. 500.00</strong>
                    </td>
                </tr>
                <tr>
                    <td style="padding:10px; border:1px solid #ddd;"><strong>Status</strong></td>
                    <td style="padding:10px; border:1px solid #ddd; color:#e67e22;">
                        <strong>PENDING</strong>
                    </td>
                </tr>
            </table>
            <p>Please log in to your student portal and pay the fine at your earliest convenience.</p>
            <p style="color:#666;">Improve your attendance to avoid future fines.</p>
            <p>Regards,<br><strong>AI Attendance System</strong></p>
        </div>
    </div>
    """
    send_email(student_info['email'], 'Attendance Fine Issued - AI Attendance System', email_body)

    cursor.close()
    conn.close()
    return jsonify({"status": "success", "message": "Fine issued successfully"})


# ============================================
# GET MY FINES (STUDENT)
# ============================================
@app.route('/my_fines', methods=['GET'])
@role_required(['Student'])
def get_my_fines():
    conn = get_db_connection()
    if conn is None:
        return jsonify({"status": "error", "message": "DB connection failed"}), 500
    
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT fine_id, course_code, course_name, 
               attendance_percentage, fine_amount, 
               status, issued_date
        FROM fines
        WHERE student_id = %s
        ORDER BY issued_date DESC
    """, (session['user_id'],))
    fines = cursor.fetchall()
    
    for fine in fines:
        if fine['issued_date']:
            fine['issued_date'] = fine['issued_date'].strftime('%Y-%m-%d')
    
    cursor.close()
    conn.close()
    return jsonify({"status": "success", "fines": fines})


# ============================================
# PAY FINE (STUDENT)
# ============================================
# ============================================
# PAY FINE (STUDENT)
# ============================================
@app.route('/pay_fine', methods=['POST'])
@role_required(['Student'])
def pay_fine():
    data = request.get_json()
    fine_id = data.get('fine_id')
    payment_intent_id = data.get('payment_intent_id', None)  # ← Add this
    
    conn = get_db_connection()
    if conn is None:
        return jsonify({"status": "error", "message": "DB connection failed"}), 500
    
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE fines 
        SET status = 'Paid', paid_date = NOW()
        WHERE fine_id = %s AND student_id = %s AND status = 'Pending'
    """, (fine_id, session['user_id']))
    conn.commit()
    
    if cursor.rowcount == 0:
        cursor.close()
        conn.close()
        return jsonify({"status": "error", "message": "Fine not found or already paid"}), 400

    # Fine info + student details fetch karo
    cursor2 = conn.cursor(dictionary=True)
    cursor2.execute("""
        SELECT u.email, u.full_name, u.user_id as roll_no,
               f.course_code, f.course_name, f.fine_amount, 
               f.attendance_percentage, f.fine_id
        FROM fines f 
        JOIN users u ON f.student_id = u.user_id 
        WHERE f.fine_id = %s
    """, (fine_id,))
    fine_info = cursor2.fetchone()
    cursor2.close()
    cursor.close()
    conn.close()

    if not fine_info:
        # Agar fine_info nahi mila to bhi receipt mat bhejo
        return jsonify({"status": "success", "message": "Fine paid successfully"})

    paid_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    receipt_no = f"REC-{datetime.now().strftime('%Y%m%d')}-{fine_info['fine_id']:04d}"

    # ---- Student ko confirmation email ----
    student_body = f"""
    <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
        <div style="background: #27ae60; padding: 20px; text-align: center;">
            <h2 style="color: white; margin: 0;">✅ Payment Successful!</h2>
        </div>
        <div style="padding: 30px; background: #f9f9f9;">
            <p>Dear <strong>{fine_info['full_name']}</strong>,</p>
            <p>Your fine payment has been successfully processed.</p>
            <table style="width:100%; border-collapse:collapse; margin:20px 0;">
                <tr style="background:#ecf0f1;">
                    <td style="padding:10px; border:1px solid #ddd;"><strong>Course</strong></td>
                    <td style="padding:10px; border:1px solid #ddd;">{fine_info['course_code']} - {fine_info['course_name']}</td>
                </tr>
                <tr>
                    <td style="padding:10px; border:1px solid #ddd;"><strong>Amount Paid</strong></td>
                    <td style="padding:10px; border:1px solid #ddd; color:#27ae60;">
                        <strong>Rs. {fine_info['fine_amount']}</strong>
                    </td>
                </tr>
                <tr style="background:#ecf0f1;">
                    <td style="padding:10px; border:1px solid #ddd;"><strong>Status</strong></td>
                    <td style="padding:10px; border:1px solid #ddd; color:#27ae60;">
                        <strong>PAID</strong>
                    </td>
                </tr>
                <tr>
                    <td style="padding:10px; border:1px solid #ddd;"><strong>Paid At</strong></td>
                    <td style="padding:10px; border:1px solid #ddd;">{paid_at}</td>
                </tr>
            </table>
            <p style="color:#666;">Please improve your attendance to avoid future fines.</p>
            <p>Regards,<br><strong>AI Attendance System</strong></p>
        </div>
    </div>
    """
    result1 = send_email(fine_info['email'], 'Fine Payment Confirmation - AI Attendance System', student_body)
    print(f"[EMAIL] Student email to {fine_info['email']}: {'OK' if result1 else 'FAILED'}")

    # ---- Admin ko notification email ----
    ADMIN_EMAIL = "aqsahashmi483@gmail.com"
    admin_body = f"""
    <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
        <div style="background: #2c3e50; padding: 20px; text-align: center;">
            <h2 style="color: white; margin: 0;">💰 Fine Payment Received</h2>
        </div>
        <div style="padding: 30px; background: #f9f9f9;">
            <p>Dear <strong>Admin</strong>,</p>
            <p>A student has successfully paid their attendance fine.</p>
            <table style="width:100%; border-collapse:collapse; margin:20px 0;">
                <tr style="background:#ecf0f1;">
                    <td style="padding:10px; border:1px solid #ddd;"><strong>Student Name</strong></td>
                    <td style="padding:10px; border:1px solid #ddd;">{fine_info['full_name']}</td>
                </tr>
                <tr>
                    <td style="padding:10px; border:1px solid #ddd;"><strong>Student Email</strong></td>
                    <td style="padding:10px; border:1px solid #ddd;">{fine_info['email']}</td>
                </tr>
                <tr style="background:#ecf0f1;">
                    <td style="padding:10px; border:1px solid #ddd;"><strong>Course</strong></td>
                    <td style="padding:10px; border:1px solid #ddd;">{fine_info['course_code']} - {fine_info['course_name']}</td>
                </tr>
                <tr>
                    <td style="padding:10px; border:1px solid #ddd;"><strong>Amount Paid</strong></td>
                    <td style="padding:10px; border:1px solid #ddd; color:#27ae60;">
                        <strong>Rs. {fine_info['fine_amount']}</strong>
                    </td>
                </tr>
                <tr style="background:#ecf0f1;">
                    <td style="padding:10px; border:1px solid #ddd;"><strong>Paid At</strong></td>
                    <td style="padding:10px; border:1px solid #ddd;">{paid_at}</td>
                </tr>
            </table>
            <p style="color:#666;">Login to Admin Dashboard to view updated fine records.</p>
            <p>Regards,<br><strong>AI Attendance System</strong></p>
        </div>
    </div>
    """
    result2 = send_email(ADMIN_EMAIL, f"Fine Paid by {fine_info['full_name']} - AI Attendance System", admin_body)
    print(f"[EMAIL] Admin notification to {ADMIN_EMAIL}: {'OK' if result2 else 'FAILED'}")

    # 🔥 RECEIPT DATA — AB YEH EXECUTE HOGA
    receipt_data = {
        "receipt_no": receipt_no,
        "student_name": fine_info['full_name'],
        "roll_no": fine_info.get('roll_no', 'N/A'),
        "course": f"{fine_info['course_code']} - {fine_info['course_name']}",
        "amount": fine_info['fine_amount'],
        "attendance": fine_info['attendance_percentage'],
        "status": "PAID",
        "paid_at": paid_at,
        "payment_id": payment_intent_id or f"pi_{datetime.now().strftime('%Y%m%d%H%M%S')}"
    }

    return jsonify({
        "status": "success", 
        "message": "Fine paid successfully",
        "receipt": receipt_data  # ← Ye frontend ko receive hoga
    })

# ============================================
# STRIPE PAYMENT INTENT (STUDENT)
# ============================================
@app.route('/create_payment_intent', methods=['POST'])
@role_required(['Student'])
def create_payment_intent():
    data = request.get_json() or {}
    fine_id = data.get('fine_id')
    amount = data.get('amount')

    if not fine_id or not amount:
        return jsonify({"status": "error", "message": "Missing fields"}), 400

    try:
        # Convert to PKR paisa (Stripe expects smallest currency unit)
        amount_in_paisa = int(float(amount)) * 100
        
        intent = stripe.PaymentIntent.create(
            amount=amount_in_paisa,
            currency="pkr",
            metadata={
                "fine_id": fine_id,
                "student_id": session.get('user_id')
            }
        )
        return jsonify({
            "status": "success",
            "client_secret": intent.client_secret
        })
    except Exception as e:
        print(f"[STRIPE ERROR] {str(e)}")
        return jsonify({"status": "error", "message": str(e)}), 500


# ============================================
# GET STRIPE PUBLISHABLE KEY (STUDENT)
# ============================================
@app.route('/get_stripe_key', methods=['GET'])
@role_required(['Student'])
def get_stripe_key():
    return jsonify({
        "publishable_key": os.getenv("STRIPE_PUBLISHABLE_KEY")
    })

# ============================================
# LOGOUT
# ============================================

@app.route('/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({"status": "success", "message": "Logged out successfully"})

# ============================================
# ADMIN APIs
# ============================================

@app.route('/admin_stats', methods=['GET'])
@role_required(['Admin'])
def admin_stats():
    conn = get_db_connection()
    if conn is None:
        return jsonify({"status": "error", "message": "DB connection failed"}), 500
    
    cursor = conn.cursor(dictionary=True)
    
    cursor.execute("SELECT COUNT(*) as total FROM users WHERE role = 'Student'")
    students = cursor.fetchone()
    
    cursor.execute("SELECT COUNT(*) as total FROM users WHERE role = 'Teacher'")
    teachers = cursor.fetchone()
    
    cursor.execute("SELECT COUNT(*) as total FROM courses")
    courses = cursor.fetchone()
    
    cursor.execute("SELECT COUNT(*) as total FROM attendance_sessions")
    sessions = cursor.fetchone()
    
    cursor.close()
    conn.close()
    
    return jsonify({
        'status': 'success',
        'total_students': students['total'] if students else 0,
        'total_teachers': teachers['total'] if teachers else 0,
        'total_courses': courses['total'] if courses else 0,
        'total_sessions': sessions['total'] if sessions else 0
    })

@app.route('/admin_attendance_trend', methods=['GET'])
@role_required(['Admin'])
def admin_attendance_trend():
    conn = get_db_connection()
    if conn is None:
        return jsonify({"status": "error", "message": "DB connection failed"}), 500

    cursor = conn.cursor(dictionary=True)
    today = date.today()
    start_date = today.fromordinal(today.toordinal() - 6)

    cursor.execute("""
        SELECT
            s.session_id,
            s.session_date,
            sec_strength.total_students,
            COALESCE(SUM(CASE WHEN UPPER(ar.status) = 'PRESENT' THEN 1 ELSE 0 END), 0) AS present_count
        FROM attendance_sessions s
        JOIN (
            SELECT section_id, COUNT(*) AS total_students
            FROM student_enrollment
            GROUP BY section_id
        ) sec_strength ON sec_strength.section_id = s.section_id
        LEFT JOIN attendance_records ar ON ar.session_id = s.session_id
        WHERE s.session_date BETWEEN %s AND %s
        GROUP BY s.session_id, s.session_date, sec_strength.total_students
        ORDER BY s.session_date ASC
    """, (start_date, today))
    session_rows = cursor.fetchall()
    cursor.close()
    conn.close()

    trend_map = {}
    for day_offset in range(7):
        day = start_date.fromordinal(start_date.toordinal() + day_offset)
        day_key = day.strftime('%Y-%m-%d')
        trend_map[day_key] = {
            "label": day.strftime('%a'),
            "present_count": 0,
            "total_students": 0
        }

    for row in session_rows:
        day_key = row['session_date'].strftime('%Y-%m-%d')
        if day_key not in trend_map:
            continue
        trend_map[day_key]["present_count"] += int(row.get('present_count') or 0)
        trend_map[day_key]["total_students"] += int(row.get('total_students') or 0)

    trend = []
    for day_key in sorted(trend_map.keys()):
        day_data = trend_map[day_key]
        total_students = day_data["total_students"]
        present_count = day_data["present_count"]
        percentage = round((present_count * 100.0 / total_students), 2) if total_students > 0 else 0
        trend.append({
            "date": day_key,
            "label": day_data["label"],
            "percentage": percentage,
            "present_count": present_count,
            "total_students": total_students
        })

    return jsonify({"status": "success", "trend": trend})

@app.route('/admin_students', methods=['GET'])
@role_required(['Admin'])
def admin_students():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT u.user_id, u.full_name, u.email, u.registration_no, u.phone, u.dept_id,
               d.dept_name, u.current_sem_id, sem.semester_number
        FROM users u
        LEFT JOIN departments d ON u.dept_id = d.dept_id
        LEFT JOIN semesters sem ON sem.sem_id = u.current_sem_id
        WHERE u.role = 'Student'
    """)
    students = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify({'status': 'success', 'students': students})

@app.route('/admin_teachers', methods=['GET'])
@role_required(['Admin'])
def admin_teachers():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT user_id, full_name, email, employee_id, qualification, phone
        FROM users WHERE role = 'Teacher'
    """)
    teachers = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify({'status': 'success', 'teachers': teachers})

@app.route('/admin_user/<int:user_id>', methods=['GET'])
@role_required(['Admin'])
def admin_get_user(user_id):
    conn = get_db_connection()
    if conn is None:
        return jsonify({"status": "error", "message": "DB connection failed"}), 500

    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT u.user_id, u.email, u.full_name, u.phone, u.role, u.registration_no,
               u.employee_id, u.qualification, u.dept_id, u.program_id, u.current_sem_id,
               d.dept_name, sem.semester_number
        FROM users u
        LEFT JOIN departments d ON u.dept_id = d.dept_id
        LEFT JOIN semesters sem ON sem.sem_id = u.current_sem_id
        WHERE u.user_id = %s
    """, (user_id,))
    row = cursor.fetchone()
    cursor.close()
    conn.close()

    if not row:
        return jsonify({"status": "error", "message": "User not found"}), 404
    if row.get("role") not in ("Student", "Teacher"):
        return jsonify({"status": "error", "message": "Cannot view this user"}), 403

    return jsonify({"status": "success", "user": row})

@app.route('/admin_user/<int:user_id>', methods=['PUT'])
@role_required(['Admin'])
def admin_update_user(user_id):
    data = request.get_json() or {}
    conn = get_db_connection()
    if conn is None:
        return jsonify({"status": "error", "message": "DB connection failed"}), 500

    cursor = conn.cursor(dictionary=True)
    cursor.execute(
        "SELECT user_id, email, role FROM users WHERE user_id = %s LIMIT 1",
        (user_id,),
    )
    existing = cursor.fetchone()
    if not existing:
        cursor.close()
        conn.close()
        return jsonify({"status": "error", "message": "User not found"}), 404
    if existing.get("role") not in ("Student", "Teacher"):
        cursor.close()
        conn.close()
        return jsonify({"status": "error", "message": "Cannot edit this user"}), 403

    full_name = (data.get("full_name") or "").strip()
    email = (data.get("email") or "").strip().lower()
    phone = (data.get("phone") or "").strip() or None
    password = data.get("password") or ""

    if not full_name or not email:
        cursor.close()
        conn.close()
        return jsonify({"status": "error", "message": "Full name and email are required"}), 400

    if email != (existing.get("email") or "").lower():
        cursor.execute("SELECT user_id FROM users WHERE email = %s LIMIT 1", (email,))
        if cursor.fetchone():
            cursor.close()
            conn.close()
            return jsonify({"status": "error", "message": "Email already in use"}), 400

    role = existing["role"]
    try:
        if role == "Student":
            registration_no = (data.get("registration_no") or "").strip()
            department_id = data.get("department_id")
            semester_number = data.get("semester_number")

            if not registration_no:
                cursor.close()
                conn.close()
                return jsonify({"status": "error", "message": "Roll number is required"}), 400
            if department_id in [None, ""]:
                cursor.close()
                conn.close()
                return jsonify({"status": "error", "message": "Department is required"}), 400
            try:
                department_id = int(department_id)
            except (TypeError, ValueError):
                cursor.close()
                conn.close()
                return jsonify({"status": "error", "message": "Invalid department"}), 400

            semester_raw = (semester_number or "").strip()
            if not semester_raw:
                cursor.close()
                conn.close()
                return jsonify({"status": "error", "message": "Semester is required"}), 400
            try:
                semester_int = int(semester_raw)
            except (TypeError, ValueError):
                cursor.close()
                conn.close()
                return jsonify({"status": "error", "message": "Invalid semester"}), 400

            cursor.execute(
                "SELECT program_id FROM programs WHERE dept_id = %s LIMIT 1",
                (department_id,),
            )
            prog = cursor.fetchone()
            if not prog:
                cursor.close()
                conn.close()
                return jsonify({"status": "error", "message": "Program not found for department"}), 400
            program_id = prog["program_id"]

            cursor.execute(
                """
                SELECT s.sem_id
                FROM semesters s
                JOIN programs p ON p.program_id = s.program_id
                WHERE p.dept_id = %s AND s.semester_number = %s
                LIMIT 1
                """,
                (department_id, semester_int),
            )
            sem_row = cursor.fetchone()
            if not sem_row:
                cursor.close()
                conn.close()
                return jsonify({"status": "error", "message": "Semester not found for department"}), 400
            current_sem_id = sem_row["sem_id"]

            if password:
                hashed = bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
                cursor.execute(
                    """
                    UPDATE users
                    SET full_name = %s, email = %s, phone = %s, registration_no = %s,
                        dept_id = %s, program_id = %s, current_sem_id = %s, password = %s
                    WHERE user_id = %s AND role = 'Student'
                    """,
                    (
                        full_name,
                        email,
                        phone,
                        registration_no,
                        department_id,
                        program_id,
                        current_sem_id,
                        hashed,
                        user_id,
                    ),
                )
            else:
                cursor.execute(
                    """
                    UPDATE users
                    SET full_name = %s, email = %s, phone = %s, registration_no = %s,
                        dept_id = %s, program_id = %s, current_sem_id = %s
                    WHERE user_id = %s AND role = 'Student'
                    """,
                    (
                        full_name,
                        email,
                        phone,
                        registration_no,
                        department_id,
                        program_id,
                        current_sem_id,
                        user_id,
                    ),
                )
        else:
            employee_id = (data.get("employee_id") or "").strip()
            qualification = (data.get("qualification") or "").strip() or None
            if not employee_id:
                cursor.close()
                conn.close()
                return jsonify({"status": "error", "message": "Employee ID is required"}), 400

            if password:
                hashed = bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
                cursor.execute(
                    """
                    UPDATE users
                    SET full_name = %s, email = %s, phone = %s, employee_id = %s,
                        qualification = %s, password = %s
                    WHERE user_id = %s AND role = 'Teacher'
                    """,
                    (full_name, email, phone, employee_id, qualification, hashed, user_id),
                )
            else:
                cursor.execute(
                    """
                    UPDATE users
                    SET full_name = %s, email = %s, phone = %s, employee_id = %s, qualification = %s
                    WHERE user_id = %s AND role = 'Teacher'
                    """,
                    (full_name, email, phone, employee_id, qualification, user_id),
                )

        if cursor.rowcount == 0:
            conn.rollback()
            cursor.close()
            conn.close()
            return jsonify({"status": "error", "message": "Update failed"}), 400

        conn.commit()
    except Exception as exc:
        conn.rollback()
        cursor.close()
        conn.close()
        if "Duplicate entry" in str(exc):
            return jsonify({"status": "error", "message": "Duplicate value (email / roll / employee id)"}), 400
        return jsonify({"status": "error", "message": "Failed to update user"}), 500

    cursor.close()
    conn.close()
    return jsonify({"status": "success", "message": "User updated"})

@app.route('/admin_user/<int:user_id>', methods=['DELETE'])
@role_required(['Admin'])
def admin_delete_user(user_id):
    if session.get("user_id") == user_id:
        return jsonify({"status": "error", "message": "You cannot delete your own account"}), 400

    conn = get_db_connection()
    if conn is None:
        return jsonify({"status": "error", "message": "DB connection failed"}), 500

    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT user_id, role FROM users WHERE user_id = %s", (user_id,))
    row = cursor.fetchone()
    if not row:
        cursor.close()
        conn.close()
        return jsonify({"status": "error", "message": "User not found"}), 404
    if row.get("role") not in ("Student", "Teacher"):
        cursor.close()
        conn.close()
        return jsonify({"status": "error", "message": "Cannot delete this user"}), 403

    role = row["role"]
    try:
        if role == "Student":
            cursor.execute("DELETE FROM attendance_records WHERE student_id = %s", (user_id,))
            cursor.execute("DELETE FROM student_enrollment WHERE student_id = %s", (user_id,))
            cursor.execute("DELETE FROM fines WHERE student_id = %s", (user_id,))
            cursor.execute("DELETE FROM facial_data WHERE student_id = %s", (user_id,))
            cursor.execute("DELETE FROM users WHERE user_id = %s AND role = 'Student'", (user_id,))
        else:
            cursor.execute(
                "SELECT section_id FROM sections WHERE teacher_id = %s",
                (user_id,),
            )
            section_rows = cursor.fetchall() or []
            section_ids = [r["section_id"] for r in section_rows]
            if section_ids:
                placeholders = ",".join(["%s"] * len(section_ids))
                cursor.execute(
                    f"DELETE FROM attendance_records WHERE section_id IN ({placeholders})",
                    tuple(section_ids),
                )
                cursor.execute(
                    f"DELETE FROM attendance_sessions WHERE section_id IN ({placeholders})",
                    tuple(section_ids),
                )
                cursor.execute(
                    f"DELETE FROM student_enrollment WHERE section_id IN ({placeholders})",
                    tuple(section_ids),
                )
                cursor.execute(
                    f"DELETE FROM sections WHERE section_id IN ({placeholders})",
                    tuple(section_ids),
                )
            cursor.execute(
                "DELETE FROM attendance_sessions WHERE teacher_id = %s",
                (user_id,),
            )
            cursor.execute("DELETE FROM users WHERE user_id = %s AND role = 'Teacher'", (user_id,))

        conn.commit()
    except Exception as exc:
        conn.rollback()
        cursor.close()
        conn.close()
        return jsonify({"status": "error", "message": str(exc)}), 500

    cursor.close()
    conn.close()
    return jsonify({"status": "success", "message": "User deleted"})

@app.route('/admin_departments', methods=['GET'])
@role_required(['Admin'])
def admin_departments():
    conn = get_db_connection()
    if conn is None:
        return jsonify({"status": "error", "message": "DB connection failed"}), 500

    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT dept_id, dept_code, dept_name
        FROM departments
        ORDER BY dept_name ASC
    """)
    departments = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify({"status": "success", "departments": departments})

def fetch_report_rows(start_date, end_date):
    conn = get_db_connection()
    if conn is None:
        return None

    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT
            s.session_date,
            c.course_code,
            c.course_name,
            sec.section_code,
            t.full_name AS teacher_name,
            st.full_name AS student_name,
            st.registration_no,
            st.email AS student_email,
            ar.status,
            ar.mode,
            ar.marked_at
        FROM attendance_records ar
        JOIN attendance_sessions s ON ar.session_id = s.session_id
        JOIN sections sec ON ar.section_id = sec.section_id
        JOIN course_semester cs ON sec.cs_id = cs.cs_id
        JOIN courses c ON cs.course_id = c.course_id
        JOIN users st ON ar.student_id = st.user_id
        JOIN users t ON s.teacher_id = t.user_id
        WHERE s.session_date BETWEEN %s AND %s
        ORDER BY s.session_date ASC, c.course_code ASC, sec.section_code ASC, st.full_name ASC
    """, (start_date, end_date))
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return rows

def build_excel_report(rows, start_date_str, end_date_str):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except Exception:
        return jsonify({"status": "error", "message": "Excel report dependency missing. Run: pip install -r requirements.txt"}), 500

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    headers = [
        "Session Date", "Course Code", "Course Name", "Section", "Teacher",
        "Student", "Roll No", "Student Email", "Status", "Mode", "Marked At"
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in rows:
        ws.append([
            row["session_date"].strftime("%Y-%m-%d") if row.get("session_date") else "",
            row.get("course_code", ""),
            row.get("course_name", ""),
            row.get("section_code", ""),
            row.get("teacher_name", ""),
            row.get("student_name", ""),
            row.get("registration_no", ""),
            row.get("student_email", ""),
            row.get("status", ""),
            row.get("mode", ""),
            row["marked_at"].strftime("%Y-%m-%d %H:%M:%S") if row.get("marked_at") else ""
        ])

    for column_cells in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 40)

    report_stream = io.BytesIO()
    wb.save(report_stream)
    report_stream.seek(0)
    filename = f"attendance_report_{start_date_str}_to_{end_date_str}.xlsx"

    return send_file(
        report_stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

def build_pdf_report(rows, start_date_str, end_date_str):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except Exception:
        return jsonify({"status": "error", "message": "PDF report dependency missing. Run: pip install -r requirements.txt"}), 500

    report_stream = io.BytesIO()
    doc = SimpleDocTemplate(report_stream, pagesize=landscape(A4), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("Attendance Report", styles["Title"]),
        Paragraph(f"Date Range: {start_date_str} to {end_date_str}", styles["Normal"]),
        Spacer(1, 10),
    ]

    table_data = [[
        "Date", "Course", "Section", "Teacher", "Student", "Roll No", "Status", "Mode", "Marked At"
    ]]
    for row in rows:
        table_data.append([
            row["session_date"].strftime("%Y-%m-%d") if row.get("session_date") else "",
            f'{row.get("course_code", "")} - {row.get("course_name", "")}',
            row.get("section_code", ""),
            row.get("teacher_name", ""),
            row.get("student_name", ""),
            row.get("registration_no", ""),
            row.get("status", ""),
            row.get("mode", ""),
            row["marked_at"].strftime("%Y-%m-%d %H:%M") if row.get("marked_at") else "",
        ])

    report_table = Table(table_data, repeatRows=1)
    report_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F9FC")]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elements.append(report_table)

    doc.build(elements)
    report_stream.seek(0)
    filename = f"attendance_report_{start_date_str}_to_{end_date_str}.pdf"
    return send_file(report_stream, as_attachment=True, download_name=filename, mimetype="application/pdf")

def _handle_admin_reports_download():
    start_date_str = request.args.get('start_date', '').strip()
    end_date_str = request.args.get('end_date', '').strip()
    format_type = request.args.get('format', 'pdf').strip().lower()

    if not start_date_str or not end_date_str:
        return jsonify({"status": "error", "message": "Start date and end date are required"}), 400

    if format_type not in ['pdf', 'excel']:
        return jsonify({"status": "error", "message": "Invalid format. Allowed: pdf, excel"}), 400

    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
    except ValueError:
        return jsonify({"status": "error", "message": "Invalid date format. Use YYYY-MM-DD"}), 400

    if start_date > end_date:
        return jsonify({"status": "error", "message": "Start date cannot be after end date"}), 400

    rows = fetch_report_rows(start_date, end_date)
    if rows is None:
        return jsonify({"status": "error", "message": "DB connection failed"}), 500

    if format_type == 'excel':
        return build_excel_report(rows, start_date_str, end_date_str)
    return build_pdf_report(rows, start_date_str, end_date_str)

@app.route('/admin_fines_list', methods=['GET'])
@role_required(['Admin'])
def admin_fines_list():
    conn = get_db_connection()
    if conn is None:
        return jsonify({"status": "error", "message": "DB connection failed"}), 500
    
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT 
            f.fine_id,
            u.full_name,
            u.email,
            f.course_code,
            f.course_name,
            f.attendance_percentage,
            f.fine_amount,
            f.status,
            DATE_FORMAT(f.issued_date, '%Y-%m-%d %H:%i') as issued_date,
            DATE_FORMAT(f.paid_date, '%Y-%m-%d %H:%i') as paid_date
        FROM fines f
        JOIN users u ON f.student_id = u.user_id
        ORDER BY 
            CASE f.status WHEN 'Pending' THEN 0 ELSE 1 END,
            f.issued_date DESC
    """)
    fines = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify({"status": "success", "fines": fines})

@app.route('/admin_reports_download', methods=['GET'])
@role_required(['Admin'])
def admin_reports_download():
    return _handle_admin_reports_download()

# Backward-compatible alias for older frontend builds.
@app.route('/generate_report', methods=['GET'])
@role_required(['Admin'])
def generate_report_legacy():
    return _handle_admin_reports_download()

@app.route('/admin_report_download', methods=['GET'])
@role_required(['Admin'])
def admin_report_download_alias():
    return _handle_admin_reports_download()

@app.route('/download_report', methods=['GET'])
@role_required(['Admin'])
def download_report_alias():
    return _handle_admin_reports_download()

# ============================================
# FACE RECOGNITION ADMIN ACTIONS
# ============================================

@app.route('/api/mark_face_attendance', methods=['POST'])
@role_required(['Student'])
def mark_face_attendance():
    data = request.get_json() or {}
    session_id = data.get('session_id')
    image_base64 = data.get('image')
    user_id = session.get('user_id')

    if not session_id:
        return jsonify({"status": "error", "message": "Session ID is required."}), 400
    if not image_base64:
        return jsonify({"status": "error", "message": "Face image is required."}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({"status": "error", "message": "Database connection failed."}), 500

    try:
        cursor = conn.cursor(dictionary=True)

        # ✅ CHECK 1: Session active aur Face/Hybrid mode mein hai?
        cursor.execute("""
            SELECT session_id, section_id, teacher_id, is_active, mode
            FROM attendance_sessions
            WHERE session_id = %s
        """, (session_id,))
        session_row = cursor.fetchone()

        if not session_row:
            return jsonify({"status": "error", "message": "Session not found. Please verify the Session ID with your teacher."}), 404

        if session_row['is_active'] != 1:
            return jsonify({"status": "error", "message": "This session has ended. Teacher has closed the attendance."}), 400

        session_mode = (session_row.get('mode') or 'Pending').strip()
        if session_mode not in ('Face', 'Hybrid'):
            return jsonify({"status": "error", "message": f"Face attendance is not active. Teacher has set mode to: {session_mode}."}), 400

        section_id = session_row['section_id']

        # ✅ CHECK 2: Student is section mein enrolled hai?
        cursor.execute("""
            SELECT enrollment_id
            FROM student_enrollment
            WHERE student_id = %s AND section_id = %s
            LIMIT 1
        """, (user_id, section_id))
        if not cursor.fetchone():
            return jsonify({"status": "error", "message": "You are not enrolled in this course/section. Face attendance denied."}), 403

        # ✅ CHECK 3: Same session mein already marked?
        cursor.execute("""
            SELECT record_id FROM attendance_records
            WHERE session_id = %s AND student_id = %s
            LIMIT 1
        """, (session_id, user_id))
        if cursor.fetchone():
            return jsonify({"status": "already_marked", "message": "Attendance already marked for this session."}), 200

        # ✅ CHECK 4: Aaj is section mein already marked?
        cursor.execute("""
            SELECT record_id FROM attendance_records
            WHERE student_id = %s AND section_id = %s AND attendance_date = CURDATE()
            LIMIT 1
        """, (user_id, section_id))
        if cursor.fetchone():
            return jsonify({"status": "already_marked", "message": "Attendance already marked for this course today."}), 200

        # ✅ CHECK 5: Face match karo (baad mein — DB checks pehle, AI baad mein)
        match_result = face_ai.match_face(user_id, image_base64)
        if match_result['status'] == 'error':
            return jsonify({
                "status": "error",
                "message": match_result.get('message', 'Face not recognized. Please try again in better lighting.')
            }), 400

        # ✅ INSERT attendance
        cursor.execute("""
            INSERT INTO attendance_records
                (session_id, student_id, section_id, attendance_date, status, mode, sync_status, marked_at)
            VALUES (%s, %s, %s, CURDATE(), 'Present', 'Face', 'Synced', NOW())
        """, (session_id, user_id, section_id))
        conn.commit()

        return jsonify({"status": "success", "message": "Attendance marked successfully via Face Recognition!"})

    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"status": "error", "message": f"Server error: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

@app.route('/api/admin/sync_faces', methods=['POST'])
@role_required(['Admin'])
def sync_faces():
    """Triggers the dataset sync process"""
    try:
        result = face_ai.sync_dataset_to_db()
        return jsonify(result)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/admin/face_stats', methods=['GET'])
@role_required(['Admin'])
def get_face_stats():
    """Returns statistics about synced faces"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Count total students with face data
        cursor.execute("SELECT COUNT(DISTINCT student_id) FROM facial_data")
        student_count = cursor.fetchone()[0]
        
        # Count total embeddings
        cursor.execute("SELECT COUNT(*) FROM facial_data")
        total_faces = cursor.fetchone()[0]
        
        return jsonify({
            "status": "success",
            "student_count": student_count,
            "total_faces": total_faces
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@app.route('/admin_add_user', methods=['POST'])
@role_required(['Admin'])
def admin_add_user():
    data = request.get_json() or {}

    full_name = (data.get('full_name') or "").strip()
    email = (data.get('email') or "").strip().lower()
    password = data.get('password') or ""
    role = (data.get('role') or "").strip()
    phone = (data.get('phone') or "").strip() or None
    department_id = data.get('department_id')

    if not full_name or not email or not password or not role:
        return jsonify({"status": "error", "message": "Full name, email, password and role are required"}), 400

    if role not in ['Student', 'Teacher']:
        return jsonify({"status": "error", "message": "Invalid role. Allowed roles: Student, Teacher"}), 400

    conn = get_db_connection()
    if conn is None:
        return jsonify({"status": "error", "message": "DB connection failed"}), 500

    cursor = conn.cursor(dictionary=True)

    if role == 'Student':
        if department_id in [None, ""]:
            cursor.close()
            conn.close()
            return jsonify({"status": "error", "message": "Department is required for Student"}), 400
        try:
            department_id = int(department_id)
        except (TypeError, ValueError):
            cursor.close()
            conn.close()
            return jsonify({"status": "error", "message": "Invalid department"}), 400

        cursor.execute("SELECT dept_id FROM departments WHERE dept_id = %s LIMIT 1", (department_id,))
        if not cursor.fetchone():
            cursor.close()
            conn.close()
            return jsonify({"status": "error", "message": "Selected department does not exist"}), 400

    cursor.execute("SELECT user_id FROM users WHERE email = %s LIMIT 1", (email,))
    if cursor.fetchone():
        cursor.close()
        conn.close()
        return jsonify({"status": "error", "message": "Email already exists"}), 400

    hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

    try:
        if role == 'Student':
            registration_no = (data.get('registration_no') or "").strip()
            if not registration_no:
                registration_no = f"STD-{datetime.now().year}-{uuid.uuid4().hex[:6].upper()}"

            cursor.execute("""
                INSERT INTO users (email, password, role, full_name, phone, registration_no, dept_id)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (email, hashed_password, role, full_name, phone, registration_no, department_id))

            created_detail = {"registration_no": registration_no, "employee_id": None}
        else:
            employee_id = (data.get('employee_id') or "").strip()
            qualification = (data.get('qualification') or "").strip() or None
            if not employee_id:
                employee_id = f"TCH-{datetime.now().year}-{uuid.uuid4().hex[:6].upper()}"

            cursor.execute("""
                INSERT INTO users (email, password, role, full_name, phone, employee_id, qualification)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (email, hashed_password, role, full_name, phone, employee_id, qualification))

            created_detail = {"registration_no": None, "employee_id": employee_id}

        conn.commit()
        user_id = cursor.lastrowid
    except Exception as exc:
        conn.rollback()
        cursor.close()
        conn.close()
        if "Duplicate entry" in str(exc):
            return jsonify({"status": "error", "message": "User details already exist (duplicate unique field)"}), 400
        return jsonify({"status": "error", "message": "Failed to create user"}), 500

    cursor.close()
    conn.close()

    return jsonify({
        "status": "success",
        "message": f"{role} added successfully",
        "user": {
            "user_id": user_id,
            "full_name": full_name,
            "email": email,
            "role": role,
            "registration_no": created_detail["registration_no"],
            "employee_id": created_detail["employee_id"]
        }
    })

@app.route('/admin_semesters', methods=['GET'])
@role_required(['Admin'])
def admin_semesters():
    dept_id = request.args.get('dept_id')
    if not dept_id:
        return jsonify({"status": "error", "message": "dept_id required"}), 400
    
    conn = get_db_connection()
    if conn is None:
        return jsonify({"status": "error", "message": "DB connection failed"}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT 
                s.sem_id,
                s.semester_number,
                s.semester_name,
                s.start_date,
                s.end_date,
                s.is_active
            FROM semesters s
            JOIN programs p ON s.program_id = p.program_id
            WHERE p.dept_id = %s
            ORDER BY s.semester_number ASC
        """, (dept_id,))
        
        semesters = cursor.fetchall()
        
        for s in semesters:
            if s.get('start_date'):
                s['start_date'] = str(s['start_date'])
            if s.get('end_date'):
                s['end_date'] = str(s['end_date'])
        
        return jsonify({"status": "success", "semesters": semesters})
        
    except Exception as e:
        print(f"ERROR in admin_semesters: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/admin_courses', methods=['GET'])
@role_required(['Admin'])
def admin_courses():
    sem_id = request.args.get('sem_id')
    if not sem_id:
        return jsonify({"status": "error", "message": "sem_id required"}), 400
    
    conn = get_db_connection()
    if conn is None:
        return jsonify({"status": "error", "message": "DB connection failed"}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT 
                c.course_id,
                c.course_code,
                c.course_name,
                c.credit_hours,
                c.course_type,
                cs.is_compulsory
            FROM course_semester cs
            JOIN courses c ON cs.course_id = c.course_id
            WHERE cs.sem_id = %s
            ORDER BY c.course_code ASC
        """, (sem_id,))
        
        courses = cursor.fetchall()
        return jsonify({"status": "success", "courses": courses})
        
    except Exception as e:
        print(f"ERROR in admin_courses: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conn.close()
        
# ============================================
# TEACHER APIs
# ============================================

def build_qr_payload(session_id):
    """Generate and persist a fresh QR code for a session."""
    now = datetime.now()
    timestamp = f"{now.timestamp():.6f}"
    payload_to_sign = f"SESSION:{session_id}:{timestamp}"
    signature = hmac.new(
        QR_SIGNING_SECRET.encode('utf-8'),
        payload_to_sign.encode('utf-8'),
        hashlib.sha256
    ).hexdigest()
    qr_data = f"{payload_to_sign}:{signature}"
    qr_img = qrcode.make(qr_data)

    buf = io.BytesIO()
    qr_img.save(buf, format="PNG")
    qr_base64 = base64.b64encode(buf.getvalue()).decode('utf-8')
    return qr_base64, now

def parse_qr_payload(qr_payload):
    """Expected payload format: SESSION:<session_id>:<qr_timestamp>:<signature>."""
    if not qr_payload or not isinstance(qr_payload, str):
        return None, None

    parts = qr_payload.split(":")
    if len(parts) != 4 or parts[0] != "SESSION":
        return None, None

    try:
        session_id = int(parts[1])
        qr_timestamp = float(parts[2])
        provided_signature = parts[3]
        payload_to_sign = f"SESSION:{session_id}:{parts[2]}"
        expected_signature = hmac.new(
            QR_SIGNING_SECRET.encode('utf-8'),
            payload_to_sign.encode('utf-8'),
            hashlib.sha256
        ).hexdigest()
        if not hmac.compare_digest(provided_signature, expected_signature):
            return None, None
        return session_id, qr_timestamp
    except (ValueError, TypeError):
        return None, None

# ============================================
# TIME-SLOT SCHEDULE HELPERS
# ============================================

def get_student_current_class(student_id):
    """
    Smart function to find student's current scheduled class.
    Returns: dict with section+session info, or None (show all)
    """
    conn = get_db_connection()
    if not conn:
        return None
    
    cursor = conn.cursor(dictionary=True)
    
    now = datetime.now()
    current_time = now.time()
    current_day = now.strftime('%A')
    
    # Weekend check
    is_weekend = current_day in ['Saturday', 'Sunday']
    
    # Check if student has any schedule today
    cursor.execute("""
        SELECT COUNT(*) as scheduled_count
        FROM student_enrollment se
        JOIN sections sec ON se.section_id = sec.section_id
        WHERE se.student_id = %s
          AND sec.class_day = %s
          AND sec.start_time IS NOT NULL
          AND sec.end_time IS NOT NULL
    """, (student_id, current_day))
    
    has_schedule = cursor.fetchone()['scheduled_count'] > 0
    
    # If weekend or no schedule → return None (show all sessions)
    if is_weekend or not has_schedule:
        cursor.close()
        conn.close()
        return None
    
    # Find current time-slot class
    cursor.execute("""
        SELECT 
            sec.section_id,
            sec.section_code,
            sec.start_time,
            sec.end_time,
            sec.room_no,
            c.course_code,
            c.course_name,
            u.full_name AS teacher_name,
            ats.session_id,
            ats.is_active,
            ats.mode,
            TIMESTAMPDIFF(MINUTE, 
                TIMESTAMP(CURDATE(), sec.end_time), 
                NOW()) AS minutes_past_end,
            TIMESTAMPDIFF(MINUTE, 
                NOW(), 
                TIMESTAMP(CURDATE(), sec.start_time)) AS minutes_before_start
        FROM student_enrollment se
        JOIN sections sec ON se.section_id = sec.section_id
        JOIN course_semester cs ON sec.cs_id = cs.cs_id
        JOIN courses c ON cs.course_id = c.course_id
        JOIN users u ON sec.teacher_id = u.user_id
        LEFT JOIN attendance_sessions ats 
            ON ats.section_id = sec.section_id 
            AND ats.is_active = 1
            AND ats.session_date = CURDATE()
        WHERE se.student_id = %s
          AND sec.class_day = %s
          AND sec.start_time IS NOT NULL
          AND sec.end_time IS NOT NULL
          AND sec.start_time <= %s
          AND sec.end_time >= %s
        ORDER BY sec.start_time ASC
        LIMIT 1
    """, (student_id, current_day, current_time, current_time))
    
    result = cursor.fetchone()
    cursor.close()
    conn.close()
    
    return result


def validate_student_class_timing(student_id, section_id):
    """
    Check if student can mark attendance in this section RIGHT NOW.
    Returns: (is_valid: bool, message: str)
    """
    conn = get_db_connection()
    if not conn:
        return False, "Database connection failed"
    
    cursor = conn.cursor(dictionary=True)
    
    # Check enrollment first
    cursor.execute("""
        SELECT 1 FROM student_enrollment
        WHERE student_id = %s AND section_id = %s
        LIMIT 1
    """, (student_id, section_id))
    
    if not cursor.fetchone():
        cursor.close()
        conn.close()
        return False, "You are not enrolled in this section"
    
    # Get schedule
    cursor.execute("""
        SELECT 
            sec.class_day,
            sec.start_time,
            sec.end_time,
            c.course_name
        FROM sections sec
        JOIN course_semester cs ON sec.cs_id = cs.cs_id
        JOIN courses c ON cs.course_id = c.course_id
        WHERE sec.section_id = %s
    """, (section_id,))
    
    schedule = cursor.fetchone()
    cursor.close()
    conn.close()
    
    # No schedule set → allow (backward compatible)
    if not schedule or not schedule.get('start_time') or not schedule.get('end_time'):
        return True, "No schedule restriction"
    
    now = datetime.now()
    current_time = now.time()
    current_day = now.strftime('%A')
    
    # Weekend → allow
    if current_day in ['Saturday', 'Sunday']:
        return True, "Weekend - no restriction"
    
    # Day check (only if schedule day is set)
    if schedule.get('class_day') and schedule['class_day'] != current_day:
        return False, f"This class is on {schedule['class_day']}, not today"
    
    # Time checks
    if current_time < schedule['start_time']:
        start_dt = datetime.combine(date.today(), schedule['start_time'])
        current_dt = datetime.combine(date.today(), current_time)
        wait_minutes = (start_dt - current_dt).seconds // 60
        if wait_minutes <= 10:
            return True, "Within grace period"  # 10 min early allowed
        return False, f"Class starts at {schedule['start_time']}. Wait {wait_minutes} min"
    
    if current_time > schedule['end_time']:
        end_dt = datetime.combine(date.today(), schedule['end_time'])
        current_dt = datetime.combine(date.today(), current_time)
        late_minutes = (current_dt - end_dt).seconds // 60
        if late_minutes <= 15:
            return True, "Within late allowance"  # 15 min late allowed
        return False, f"Class ended at {schedule['end_time']}. Too late"
    
    return True, "Valid time slot"


def check_teacher_overlap(teacher_id, current_section_id):
    """
    Check if teacher already has active session in another section.
    """
    conn = get_db_connection()
    if not conn:
        return {"has_overlap": False}
    
    cursor = conn.cursor(dictionary=True)
    
    cursor.execute("""
        SELECT 
            ats.session_id,
            sec.section_code,
            c.course_name,
            c.course_code
        FROM attendance_sessions ats
        JOIN sections sec ON ats.section_id = sec.section_id
        JOIN course_semester cs ON sec.cs_id = cs.cs_id
        JOIN courses c ON cs.course_id = c.course_id
        WHERE ats.teacher_id = %s
          AND ats.is_active = 1
          AND ats.section_id != %s
          AND ats.session_date = CURDATE()
        LIMIT 1
    """, (teacher_id, current_section_id))
    
    overlap = cursor.fetchone()
    cursor.close()
    conn.close()
    
    if overlap:
        return {
            "has_overlap": True,
            "message": f"Already active: {overlap['course_code']} ({overlap['section_code']})",
            "active_session_id": overlap['session_id']
        }
    
    return {"has_overlap": False}        

@app.route('/teacher_courses', methods=['GET'])
@role_required(['Teacher'])
def get_teacher_courses():
    teacher_id = session.get('user_id')
    
    conn = get_db_connection()
    if conn is None:
        return jsonify({"status": "error", "message": "DB connection failed"}), 500
    
    cursor = conn.cursor(dictionary=True)
    
    cursor.execute("""
        SELECT 
            s.section_id,
            s.section_code,
            s.room_no,
            c.course_id,
            c.course_code,
            c.course_name,
            c.credit_hours,
            dep.dept_name
        FROM sections s
        JOIN course_semester cs ON s.cs_id = cs.cs_id
        JOIN courses c ON cs.course_id = c.course_id
        JOIN departments dep ON c.dept_id = dep.dept_id
        WHERE s.teacher_id = %s AND s.is_active = 1
    """, (teacher_id,))
    
    courses = cursor.fetchall()
    cursor.close()
    conn.close()
    
    return jsonify({"status": "success", "courses": courses})

# ============================================
# SESSION MANAGEMENT (TEACHER)
# ============================================

@app.route('/update_session_mode', methods=['POST'])
@role_required(['Teacher'])
def update_session_mode():
    data = request.json
    session_id = data.get('session_id')
    mode = data.get('mode')
    
    if not session_id or not mode:
        return jsonify({"status": "error", "message": "Missing session_id or mode"}), 400
        
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE attendance_sessions SET mode = %s WHERE session_id = %s",
            (mode, session_id)
        )
        conn.commit()
        return jsonify({"status": "success", "message": f"Mode updated to {mode}"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@app.route('/start_session', methods=['POST'])
@role_required(['Teacher'])
def start_session():
    data = request.json
    teacher_id = session.get('user_id')
    section_id = data.get('section_id')

    if not section_id:
        return jsonify({"status": "error", "message": "Section ID missing"}), 400

    conn = get_db_connection()
    if conn is None:
        return jsonify({"status": "error", "message": "DB connection failed"}), 500

    cursor = conn.cursor(dictionary=True)
    cursor.execute(
        "SELECT section_id FROM sections WHERE section_id = %s AND teacher_id = %s AND is_active = 1",
        (section_id, teacher_id)
    )
        # Check teacher doesn't have another active session
    overlap = check_teacher_overlap(teacher_id, section_id)
    if overlap.get('has_overlap'):
        cursor.close()
        conn.close()
        return jsonify({
            "status": "error",
            "message": overlap['message'],
            "suggestion": "Close the existing session first before starting a new one."
        }), 400
    
    section = cursor.fetchone()
    if not section:
        cursor.close()
        conn.close()
        return jsonify({"status": "error", "message": "Section not assigned to this teacher"}), 403

    today = date.today()
    now = datetime.now()
    session_token = str(uuid.uuid4())

    cursor.execute(
        """SELECT session_id FROM attendance_sessions 
           WHERE section_id = %s AND session_date = %s AND is_active = 1""",
        (section_id, today)
    )
    existing = cursor.fetchone()
    
    if existing:
        session_id = existing['session_id']
    else:
        cursor.execute(
            """INSERT INTO attendance_sessions 
               (section_id, teacher_id, session_date, start_time, session_token, mode, is_active) 
               VALUES (%s, %s, %s, %s, %s, 'Pending', 1)""",
            (section_id, teacher_id, today, now, session_token)
        )
        session_id = cursor.lastrowid
        conn.commit()

    cursor.close()
    conn.close()

    return jsonify({
        "status": "success",
        "session_id": session_id
    })

@app.route('/generate_qr', methods=['POST'])
@role_required(['Teacher'])
def generate_qr():
    data = request.json or {}
    session_id = data.get('session_id')
    teacher_id = session.get('user_id')

    if not session_id:
        return jsonify({"status": "error", "message": "Session ID missing"}), 400

    conn = get_db_connection()
    if conn is None:
        return jsonify({"status": "error", "message": "DB connection failed"}), 500

    cursor = conn.cursor(dictionary=True)
    
    cursor.execute(
        """SELECT session_id FROM attendance_sessions
           WHERE session_id = %s AND teacher_id = %s AND is_active = 1""",
        (session_id, teacher_id)
    )
    session_row = cursor.fetchone()

    if not session_row:
        cursor.close()
        conn.close()
        return jsonify({"status": "error", "message": "Active session not found"}), 404

    qr_base64, qr_generated_at = build_qr_payload(session_id)
    
    cursor.execute(
        """UPDATE attendance_sessions 
           SET qr_code = %s, qr_generated_at = %s, mode = 'QR'
           WHERE session_id = %s""",
        (qr_base64, qr_generated_at, session_id)
    )
    conn.commit()

    cursor.close()
    conn.close()

    return jsonify({
        "status": "success",
        "session_id": session_id,
        "qr_code": qr_base64,
        "expires_in": QR_EXPIRY_SECONDS
    })

@app.route('/refresh_qr', methods=['POST'])
@role_required(['Teacher'])
def refresh_qr():
    return generate_qr()

@app.route('/stop_qr', methods=['POST'])
@role_required(['Teacher'])
def stop_qr():
    data = request.json or {}
    session_id = data.get('session_id')
    teacher_id = session.get('user_id')

    if not session_id:
        return jsonify({"status": "error", "message": "Session ID missing"}), 400

    conn = get_db_connection()
    if conn is None:
        return jsonify({"status": "error", "message": "DB connection failed"}), 500

    cursor = conn.cursor()
    
    cursor.execute(
        """UPDATE attendance_sessions
           SET qr_code = NULL, qr_generated_at = NULL, mode = 'Pending'
           WHERE session_id = %s AND teacher_id = %s AND is_active = 1""",
        (session_id, teacher_id)
    )
    conn.commit()

    if cursor.rowcount == 0:
        cursor.close()
        conn.close()
        return jsonify({"status": "error", "message": "Active session not found"}), 404

    cursor.close()
    conn.close()
    return jsonify({"status": "success", "message": "QR stopped successfully"})

@app.route('/attendance_list', methods=['GET'])
@role_required(['Teacher'])
def get_attendance_list():
    session_id = request.args.get('session_id')
    teacher_id = session.get('user_id')
    
    if not session_id:
        return jsonify({"status": "error", "message": "Session ID missing"}), 400

    conn = get_db_connection()
    if conn is None:
        return jsonify({"status": "error", "message": "DB connection failed"}), 500

    cursor = conn.cursor(dictionary=True)
    cursor.execute(
        "SELECT section_id, is_active FROM attendance_sessions WHERE session_id = %s AND teacher_id = %s",
        (session_id, teacher_id)
    )
    session_row = cursor.fetchone()
    if not session_row:
        cursor.close()
        conn.close()
        return jsonify({"status": "error", "message": "Session not found"}), 404
    
    cursor.execute("""
        SELECT 
            u.user_id,
            u.full_name as student_name,
            u.registration_no,
            CASE
                WHEN ar.record_id IS NULL THEN CASE WHEN %s = 1 THEN 'not_marked' ELSE 'absent' END
                WHEN UPPER(ar.status) = 'PRESENT' THEN 'present'
                WHEN UPPER(ar.status) = 'ABSENT' THEN 'absent'
                ELSE 'not_marked'
            END as status,
            ar.mode,
            ar.marked_at
        FROM student_enrollment se
        JOIN users u ON se.student_id = u.user_id
        LEFT JOIN attendance_records ar ON ar.student_id = u.user_id AND ar.session_id = %s
        WHERE se.section_id = %s
        ORDER BY u.full_name
    """, (session_row.get('is_active', 0), session_id, session_row['section_id']))
    
    records = cursor.fetchall()
    cursor.close()
    conn.close()

    # Convert datetime objects to ISO format strings for proper timezone handling
    for record in records:
        if record['marked_at']:
            record['marked_at'] = record['marked_at'].isoformat()
        else:
            record['marked_at'] = None

    return jsonify({"status": "success", "attendance": records})

@app.route('/close_session', methods=['POST'])
@role_required(['Teacher'])
def close_session():
    data = request.json or {}
    session_id = data.get('session_id')
    teacher_id = session.get('user_id')
    
    if not session_id:
        return jsonify({"status": "error", "message": "Session ID missing"}), 400
    
    conn = get_db_connection()
    if conn is None:
        return jsonify({"status": "error", "message": "DB connection failed"}), 500
    
    cursor = conn.cursor()
    
    cursor.execute(
        "UPDATE attendance_sessions SET is_active = 0, end_time = %s WHERE session_id = %s AND teacher_id = %s",
        (datetime.now(), session_id, teacher_id)
    )
    conn.commit()
    if cursor.rowcount == 0:
        cursor.close()
        conn.close()
        return jsonify({"status": "error", "message": "Session not found"}), 404
    
    cursor.close()
    conn.close()
    
    return jsonify({"status": "success", "message": "Session closed successfully"})

@app.route('/teacher_mark_attendance', methods=['POST'])
@role_required(['Teacher'])
def teacher_mark_attendance():
    data = request.get_json() or {}
    teacher_id = session.get('user_id')
    session_id = data.get('session_id')
    student_ids = data.get('student_ids') or []
    status = (data.get('status') or '').strip()

    if not session_id:
        return jsonify({"status": "error", "message": "Session ID missing"}), 400
    if not isinstance(student_ids, list) or not student_ids:
        return jsonify({"status": "error", "message": "student_ids must be a non-empty list"}), 400
    if status not in ["Present", "Absent"]:
        return jsonify({"status": "error", "message": "Invalid status. Allowed: Present, Absent"}), 400

    # Deduplicate and validate ids
    try:
        student_ids = sorted({int(sid) for sid in student_ids})
    except Exception:
        return jsonify({"status": "error", "message": "Invalid student_ids"}), 400

    conn = get_db_connection()
    if conn is None:
        return jsonify({"status": "error", "message": "DB connection failed"}), 500

    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute(
            """SELECT session_id, section_id, teacher_id, is_active
               FROM attendance_sessions
               WHERE session_id = %s""",
            (session_id,)
        )
        sess = cursor.fetchone()
        if not sess:
            return jsonify({"status": "error", "message": "Session not found"}), 404
        if int(sess.get('teacher_id')) != int(teacher_id):
            return jsonify({"status": "error", "message": "Access denied"}), 403
        if sess.get('is_active') != 1:
            return jsonify({"status": "error", "message": "Session is not active"}), 400

        section_id = sess.get('section_id')

        placeholders = ",".join(["%s"] * len(student_ids))
        cursor.execute(
            f"""SELECT student_id
                FROM student_enrollment
                WHERE section_id = %s AND student_id IN ({placeholders})""",
            (section_id, *student_ids)
        )
        enrolled = {row["student_id"] for row in cursor.fetchall()}
        missing = [sid for sid in student_ids if sid not in enrolled]
        if missing:
            return jsonify({"status": "error", "message": "One or more students are not enrolled in this section"}), 400

        affected = 0
        if status == "Present":
            # Upsert attendance records marked by teacher.
            for sid in student_ids:
                cursor.execute(
                    """INSERT INTO attendance_records
                       (session_id, student_id, section_id, attendance_date, mode, status, sync_status, marked_at)
                       VALUES (%s, %s, %s, CURDATE(), 'Teacher', 'Present', 'Synced', CURRENT_TIMESTAMP)
                       ON DUPLICATE KEY UPDATE
                         status = VALUES(status),
                         mode = VALUES(mode),
                         sync_status = VALUES(sync_status),
                         marked_at = CURRENT_TIMESTAMP""",
                    (session_id, sid, section_id)
                )
                affected += cursor.rowcount
        else:
            # Explicitly store Absent during active session for audit + correct UX.
            for sid in student_ids:
                cursor.execute(
                    """INSERT INTO attendance_records
                       (session_id, student_id, section_id, attendance_date, mode, status, sync_status, marked_at)
                       VALUES (%s, %s, %s, CURDATE(), 'Teacher', 'Absent', 'Synced', CURRENT_TIMESTAMP)
                       ON DUPLICATE KEY UPDATE
                         status = VALUES(status),
                         mode = VALUES(mode),
                         sync_status = VALUES(sync_status),
                         marked_at = CURRENT_TIMESTAMP""",
                    (session_id, sid, section_id)
                )
                affected += cursor.rowcount

        conn.commit()
        return jsonify({"status": "success", "updated": affected})
    except Exception as exc:
        conn.rollback()
        return jsonify({"status": "error", "message": f"Failed to update attendance: {exc}"}), 500
    finally:
        cursor.close()
        conn.close()


def _format_dt_pkt(dt_value):
    """Format datetime as PK time string for downloads."""
    if not dt_value:
        return ""
    try:
        from zoneinfo import ZoneInfo
        pkt = ZoneInfo("Asia/Karachi")
        if getattr(dt_value, "tzinfo", None) is None:
            dt_value = dt_value.replace(tzinfo=timezone.utc)
        return dt_value.astimezone(pkt).strftime("%Y-%m-%d %I:%M:%S %p")
    except Exception:
        # Fallback: return naive datetime as string.
        try:
            return dt_value.strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            return str(dt_value)


@app.route('/teacher_session_report', methods=['GET'])
@role_required(['Teacher'])
def teacher_session_report():
    teacher_id = session.get('user_id')
    session_id = request.args.get('session_id', '').strip()
    format_type = request.args.get('format', 'excel').strip().lower()

    if not session_id:
        return jsonify({"status": "error", "message": "Session ID is required"}), 400
    if format_type not in ['excel', 'pdf']:
        return jsonify({"status": "error", "message": "Invalid format. Allowed: excel, pdf"}), 400

    try:
        session_id_int = int(session_id)
    except ValueError:
        return jsonify({"status": "error", "message": "Invalid session_id"}), 400

    conn = get_db_connection()
    if conn is None:
        return jsonify({"status": "error", "message": "DB connection failed"}), 500

    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute(
            """SELECT s.session_id, s.session_date, s.section_id, s.teacher_id, s.is_active,
                      sec.section_code, c.course_code, c.course_name, u.full_name AS teacher_name
               FROM attendance_sessions s
               JOIN sections sec ON s.section_id = sec.section_id
               JOIN course_semester cs ON sec.cs_id = cs.cs_id
               JOIN courses c ON cs.course_id = c.course_id
               JOIN users u ON s.teacher_id = u.user_id
               WHERE s.session_id = %s""",
            (session_id_int,)
        )
        sess = cursor.fetchone()
        if not sess:
            return jsonify({"status": "error", "message": "Session not found"}), 404
        if int(sess.get('teacher_id')) != int(teacher_id):
            return jsonify({"status": "error", "message": "Access denied"}), 403

        cursor.execute(
            """SELECT 
                    stu.full_name AS student_name,
                    stu.registration_no,
                    stu.email AS student_email,
                    CASE
                        WHEN ar.record_id IS NULL THEN CASE WHEN %s = 1 THEN 'Not Marked' ELSE 'Absent' END
                        WHEN UPPER(ar.status) = 'PRESENT' THEN 'Present'
                        WHEN UPPER(ar.status) = 'ABSENT' THEN 'Absent'
                        ELSE 'Not Marked'
                    END AS status,
                    ar.mode,
                    ar.marked_at
               FROM student_enrollment se
               JOIN users stu ON se.student_id = stu.user_id
               LEFT JOIN attendance_records ar ON ar.student_id = stu.user_id AND ar.session_id = %s
               WHERE se.section_id = %s
               ORDER BY stu.full_name""",
            (sess.get('is_active', 0), session_id_int, sess.get('section_id'))
        )
        rows = cursor.fetchall()

        if format_type == 'pdf':
            # Prepare rows for build_pdf_report
            pdf_rows = []
            for r in rows:
                pdf_rows.append({
                    "session_date": sess.get("session_date"),
                    "course_code": sess.get("course_code"),
                    "course_name": sess.get("course_name"),
                    "section_code": sess.get("section_code"),
                    "teacher_name": sess.get("teacher_name"),
                    "student_name": r.get("student_name"),
                    "registration_no": r.get("registration_no"),
                    "student_email": r.get("student_email"),
                    "status": r.get("status"),
                    "mode": r.get("mode"),
                    "marked_at": r.get("marked_at")
                })
            date_str = sess.get("session_date").strftime("%Y-%m-%d") if sess.get("session_date") else "unknown"
            return build_pdf_report(pdf_rows, date_str, date_str)

        # Build Excel (reuse openpyxl dependency style)
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except Exception:
            return jsonify({"status": "error", "message": "Excel report dependency missing. Run: pip install -r requirements.txt"}), 500

        wb = Workbook()
        ws = wb.active
        ws.title = "Session Sheet"

        headers = [
            "Session Date", "Course Code", "Course Name", "Section", "Teacher",
            "Student", "Roll No", "Student Email", "Status", "Method", "Marked At (PKT)"
        ]
        ws.append(headers)

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        def _method_label(mode_val):
            if not mode_val:
                return ""
            m = str(mode_val).lower()
            if m == "qr":
                return "QR Scan"
            if m == "teacher" or m == "manual":
                return "Teacher Marked"
            if m == "student":
                return "Student Self"
            return str(mode_val)

        session_date = sess.get("session_date")
        for r in rows:
            ws.append([
                session_date.strftime("%Y-%m-%d") if session_date else "",
                sess.get("course_code", ""),
                sess.get("course_name", ""),
                sess.get("section_code", ""),
                sess.get("teacher_name", ""),
                r.get("student_name", ""),
                r.get("registration_no", ""),
                r.get("student_email", ""),
                r.get("status", ""),
                _method_label(r.get("mode")),
                _format_dt_pkt(r.get("marked_at")),
            ])

        for column_cells in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 40)

        report_stream = io.BytesIO()
        wb.save(report_stream)
        report_stream.seek(0)

        filename = f"attendance_session_{sess.get('course_code','course')}_{sess.get('section_code','sec')}_{session_id_int}.xlsx"
        return send_file(
            report_stream,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    finally:
        cursor.close()
        conn.close()

# ============================================
# STUDENT APIs
# ============================================

@app.route('/mark_attendance', methods=['POST'])
@role_required(['Student'])
def mark_attendance():
    data = request.json or {}
    student_id = session.get('user_id')
    mode = data.get('mode', 'QR')
    session_id = data.get('session_id')
    qr_payload = data.get('qr_payload')

    if mode == 'QR':
        parsed_session_id, qr_timestamp = parse_qr_payload(qr_payload)
        if not parsed_session_id or qr_timestamp is None:
            return jsonify({"status": "error", "message": "Invalid or tampered QR code. Please scan again."}), 400
        session_id = parsed_session_id
    else:
        qr_timestamp = None

    if not session_id:
        return jsonify({"status": "error", "message": "Session ID missing"}), 400

    conn = get_db_connection()
    if conn is None:
        return jsonify({"status": "error", "message": "Database connection failed"}), 500

    cursor = conn.cursor(dictionary=True)

    # ✅ CHECK 1: Session exists aur active hai?
    cursor.execute("""
        SELECT session_id, section_id, is_active, qr_generated_at, mode
        FROM attendance_sessions
        WHERE session_id = %s
    """, (session_id,))
    session_result = cursor.fetchone()

    if not session_result:
        cursor.close()
        conn.close()
        return jsonify({"status": "error", "message": "Session not found. Please ask your teacher to verify the session ID."}), 404

    if session_result['is_active'] != 1:
        cursor.close()
        conn.close()
        return jsonify({"status": "error", "message": "This session has ended. Teacher has closed the attendance session."}), 400

    session_mode = (session_result.get('mode') or 'Pending').strip()

    # ✅ CHECK 2: Mode validation
    if mode == 'QR':
        if session_mode != 'QR':
            cursor.close()
            conn.close()
            return jsonify({"status": "error", "message": f"QR attendance is not active. Teacher has set mode to: {session_mode}. Please use the correct method."}), 400

        # QR expiry check
        qr_generated_at = session_result.get('qr_generated_at')
        if not qr_generated_at:
            cursor.close()
            conn.close()
            return jsonify({"status": "error", "message": "QR code is not active. Please ask your teacher to generate QR."}), 400

        if datetime.now().timestamp() - qr_timestamp > QR_EXPIRY_SECONDS:
            cursor.close()
            conn.close()
            return jsonify({"status": "error", "message": "QR code has expired. Please scan the latest QR code shown by your teacher."}), 400

    elif mode == 'Manual':
        if session_mode == 'Pending':
            cursor.close()
            conn.close()
            return jsonify({"status": "error", "message": "Teacher has not started attendance yet. Please wait."}), 400
        if session_mode == 'Face':
            cursor.close()
            conn.close()
            return jsonify({"status": "error", "message": "This session uses Face Recognition. Please use face attendance method."}), 400

    section_id = session_result['section_id']

    # ✅ CHECK 3: Student is section mein enrolled hai?
    cursor.execute("""
        SELECT enrollment_id
        FROM student_enrollment
        WHERE student_id = %s AND section_id = %s
        LIMIT 1
    """, (student_id, section_id))
    if not cursor.fetchone():
        cursor.close()
        conn.close()
        return jsonify({"status": "error", "message": "You are not enrolled in this course/section. Please contact your teacher or admin."}), 403
    
        # ✅ CHECK 3.5: Time-slot validation
    is_valid_time, time_msg = validate_student_class_timing(student_id, section_id)
    
    if not is_valid_time:
        cursor.close()
        conn.close()
        return jsonify({
            "status": "error", 
            "message": time_msg
        }), 400

    # ✅ CHECK 4: Same session mein already marked?
    cursor.execute(
        "SELECT record_id FROM attendance_records WHERE session_id = %s AND student_id = %s LIMIT 1",
        (session_id, student_id)
    )
    if cursor.fetchone():
        cursor.close()
        conn.close()
        return jsonify({"status": "already_marked", "message": "Attendance already marked for this session."}), 200

    # ✅ CHECK 5: Aaj is section mein kisi bhi session mein already marked?
    cursor.execute(
        "SELECT record_id FROM attendance_records WHERE student_id = %s AND section_id = %s AND attendance_date = CURDATE() LIMIT 1",
        (student_id, section_id)
    )
    if cursor.fetchone():
        cursor.close()
        conn.close()
        return jsonify({"status": "already_marked", "message": "Attendance already marked for this course today."}), 200

    # ✅ INSERT attendance
    try:
        cursor.execute("""
            INSERT INTO attendance_records
                (session_id, student_id, section_id, attendance_date, mode, status, sync_status, marked_at)
            VALUES
                (%s, %s, %s, CURDATE(), %s, 'Present', 'Synced', NOW())
        """, (session_id, student_id, section_id, mode))
        conn.commit()
    except Exception as exc:
        conn.rollback()
        error_text = str(exc)
        if "Duplicate entry" in error_text:
            cursor.close()
            conn.close()
            return jsonify({"status": "already_marked", "message": "Attendance already marked."}), 200
        cursor.close()
        conn.close()
        return jsonify({"status": "error", "message": f"Failed to mark attendance: {error_text}"}), 500

    cursor.close()
    conn.close()
    return jsonify({"status": "success", "message": "Attendance marked successfully!"})

@app.route('/my_attendance', methods=['GET'])
@role_required(['Student'])
def get_my_attendance():
    student_id = session.get('user_id')
    
    conn = get_db_connection()
    if conn is None:
        return jsonify({"status": "error", "message": "DB connection failed"}), 500
    
    cursor = conn.cursor(dictionary=True)
    
    cursor.execute("""
        SELECT
            c.course_code,
            c.course_name,
            t.full_name AS teacher_name,
            COUNT(DISTINCT CASE WHEN UPPER(ar.status) = 'PRESENT' THEN ar.record_id END) as present_days,
            COUNT(DISTINCT ar.record_id) as total_sessions,
            ROUND(
                COUNT(DISTINCT CASE WHEN UPPER(ar.status) = 'PRESENT' THEN ar.record_id END) * 100.0 /
                NULLIF(COUNT(DISTINCT ar.record_id), 0), 2
            ) as percentage
        FROM attendance_records ar
        JOIN sections sec ON ar.section_id = sec.section_id
        JOIN course_semester cs ON sec.cs_id = cs.cs_id
        JOIN courses c ON cs.course_id = c.course_id
        JOIN attendance_sessions ats ON ar.session_id = ats.session_id
        JOIN users t ON ats.teacher_id = t.user_id
        WHERE ar.student_id = %s
        GROUP BY c.course_code, c.course_name, t.full_name
    """, (student_id,))
    
    attendance = cursor.fetchall()
    cursor.close()
    conn.close()
    
    return jsonify({"status": "success", "attendance": attendance})

# ============================================
# ACTIVE SESSIONS FOR STUDENT (NEW ENDPOINT)
# Add this in app.py after /my_attendance route
# ============================================

@app.route('/active_sessions_for_student', methods=['GET'])
@role_required(['Student'])
def active_sessions_for_student():
    student_id = session.get('user_id')

    conn = get_db_connection()
    if conn is None:
        return jsonify({"status": "error", "message": "DB connection failed"}), 500

    try:
        cursor = conn.cursor(dictionary=True)
        today_date = date.today()
        
        # Get current scheduled class
        current_class = get_student_current_class(student_id)
        
        sessions = []
        current_class_info = None
        
        if current_class is not None:
            # Schedule exists → Show only current class session
            if current_class.get('session_id'):
                # Check if already marked
                cursor.execute("""
                    SELECT 1 FROM attendance_records
                    WHERE session_id = %s AND student_id = %s
                    LIMIT 1
                """, (current_class['session_id'], student_id))
                already_marked = cursor.fetchone() is not None
                
                sessions = [{
                    'session_id': current_class['session_id'],
                    'section_id': current_class['section_id'],
                    'mode': current_class.get('mode', 'Pending'),
                    'course_code': current_class['course_code'],
                    'course_name': current_class['course_name'],
                    'section_code': current_class['section_code'],
                    'room_no': current_class.get('room_no'),
                    'teacher_name': current_class['teacher_name'],
                    'already_marked': already_marked
                }]
            
            current_class_info = {
                "course_code": current_class['course_code'],
                "course_name": current_class['course_name'],
                "teacher": current_class['teacher_name'],
                "timing": f"{current_class['start_time']} - {current_class['end_time']}"
            }
        else:
            # No schedule → Show ALL active sessions (backward compatible)
            cursor.execute("""
                SELECT
                    ats.session_id,
                    ats.section_id,
                    ats.mode,
                    c.course_code,
                    c.course_name,
                    sec.section_code,
                    sec.room_no,
                    t.full_name AS teacher_name,
                    CASE 
                        WHEN ar.record_id IS NOT NULL THEN 1 
                        ELSE 0 
                    END AS already_marked
                FROM attendance_sessions ats
                JOIN sections sec ON ats.section_id = sec.section_id
                JOIN course_semester cs ON sec.cs_id = cs.cs_id
                JOIN courses c ON cs.course_id = c.course_id
                JOIN users t ON ats.teacher_id = t.user_id
                JOIN student_enrollment se ON se.section_id = sec.section_id
                LEFT JOIN attendance_records ar 
                ON ar.section_id = ats.section_id
                AND ar.student_id = %s
                AND ar.attendance_date = CURDATE()
                WHERE ats.is_active = 1
                  AND se.student_id = %s
                  AND ats.session_date = %s
            """, (student_id, student_id, today_date))
            
            sessions = cursor.fetchall()

        return jsonify({
            "status": "success", 
            "sessions": sessions,
            "current_class": current_class_info
        })

    except Exception as e:
        print(f"ERROR in active_sessions_for_student: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@app.route('/my_attendance_history', methods=['GET'])
@role_required(['Student'])
def get_my_attendance_history():
    student_id = session.get('user_id')
    
    try:
        conn = get_db_connection()
        if conn is None:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500
        
        cursor = conn.cursor(dictionary=True)
        
        # Fetch last 10 attendance records
        cursor.execute("""
            SELECT 
                ar.record_id,
                c.course_code,
                c.course_name,
                DATE_FORMAT(asess.session_date, '%Y-%m-%d') as session_date,
                ar.status,
                ar.mode,
                ar.marked_at
            FROM attendance_records ar
            JOIN attendance_sessions asess ON ar.session_id = asess.session_id
            JOIN sections sec ON ar.section_id = sec.section_id
            JOIN course_semester cs ON sec.cs_id = cs.cs_id
            JOIN courses c ON cs.course_id = c.course_id
            WHERE ar.student_id = %s
            ORDER BY asess.session_date DESC, ar.marked_at DESC
            LIMIT 10
        """, (student_id,))
        
        history = cursor.fetchall()

        # Convert datetime objects to ISO format strings for proper timezone handling
        for record in history:
            if record['marked_at']:
                record['marked_at'] = record['marked_at'].isoformat()
            else:
                record['marked_at'] = None
        
        cursor.close()
        conn.close()
        
        return jsonify({"status": "success", "history": history})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

# ============================================
# CHATBOT API (ROLE-BASED)
# ============================================

import requests as http_requests

OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY")
CHAT_SYSTEM_PROMPTS = {
    "Student": """You are a friendly AI assistant built into an AI Attendance System for students.

Your ONLY job is to help students with topics related to:
- Their attendance, courses, marks, and academic performance
- Checking attendance percentage per course
- The 75% minimum attendance rule and its consequences
- How QR code attendance works (student scans QR shown by teacher in class)
- How Face Recognition attendance works (student uses camera to mark attendance)
- Attendance fines: Rs. 500 issued when attendance drops below 75%, paid through the student portal
- How to view fine status (Pending/Paid) in the dashboard
- Study tips, time management, exam preparation, and academic advice
- How to use the student dashboard features

STRICT RULES:
1. ALWAYS respond in English only. Never use Roman Urdu, Urdu, or any other language.
2. ONLY answer questions related to the attendance system, academics, study tips, or student life.
3. If asked about mobile phones, games, movies, politics, sports, cooking, or ANY unrelated topic, politely say: "I'm here to help with attendance and academic matters only. Is there anything related to your studies or attendance that I can assist with?"
4. Never reveal other students' personal data.
5. Never say "I cannot help" without offering an alternative academic/attendance topic.
6. Keep responses concise and helpful. Use simple English.""",

    "Teacher": """You are a helpful AI assistant built into an AI Attendance System for teachers.

Your ONLY job is to help teachers with:
- Starting and managing attendance sessions
- How QR code sessions work: QR is generated and auto-refreshes every 15 seconds, students scan to mark Present
- How Face Recognition mode works for attendance
- How to manually mark students as Present or Absent
- How to close a session once class is done
- How to download Excel reports for any session
- Viewing assigned sections and course details
- Classroom management and teaching strategies
- Best practices for maintaining student attendance records

STRICT RULES:
1. ALWAYS respond in English only. Never use Roman Urdu, Urdu, or any other language.
2. ONLY answer questions related to the attendance system, teaching, classroom management, or academic matters.
3. If asked about mobile phones, gadgets, entertainment, politics, or ANY unrelated topic, politely say: "I'm here to assist with attendance and teaching-related matters only. How can I help you with your classes or attendance sessions?"
4. If asked about admin-only features (like issuing fines, adding users), explain those are admin functions and suggest contacting the admin.
5. Never flatly refuse. Always redirect to attendance or teaching topics.
6. Keep responses clear and professional.""",

    "Admin": """You are a knowledgeable AI assistant built into an AI Attendance System for administrators.

Your ONLY job is to help administrators with:
- Adding new students and teachers
- Managing departments
- Viewing all students and teachers with their details
- Low attendance monitoring: students below 75% appear in a list, admin can issue Rs. 500 fines
- Fine management: viewing all fines (Pending/Paid), automatic email notifications
- Downloading attendance reports as PDF or Excel filtered by date range
- Attendance trends: 7-day daily attendance chart
- Face recognition sync: syncing face dataset from files to database
- QR attendance sessions overview
- System configuration via environment variables
- Server management and technical issues related to this attendance system

STRICT RULES:
1. ALWAYS respond in English only. Never use Roman Urdu, Urdu, or any other language.
2. ONLY answer questions related to the attendance system, its administration, server management, or technical support for THIS system.
3. If asked about unrelated software, mobile phones, entertainment, politics, or ANY off-topic subject, politely say: "I'm here to assist with the AI Attendance System only. Is there anything related to managing the attendance system that I can help with?"
4. Never refuse without offering useful guidance about the attendance system.
5. Be concise but complete in responses."""
}

@app.route('/chatbot', methods=['POST'])
@role_required(['Admin', 'Teacher', 'Student'])
def chatbot():
    role = session.get('role')  # 'Admin', 'Teacher', or 'Student'
    data = request.get_json() or {}
    
    user_message = (data.get('message') or '').strip()
    history = data.get('history') or []  # list of {role, content} dicts
    
    if not user_message:
        return jsonify({"status": "error", "message": "Message is required"}), 400
    
    messages = []
    for entry in history[-10:]:
        if entry.get('role') in ['user', 'assistant'] and entry.get('content'):
            messages.append({
                "role": entry['role'],
                "content": str(entry['content'])
            })
    messages.append({"role": "user", "content": user_message})
    
    try:
        or_messages = [{"role": "system", "content": CHAT_SYSTEM_PROMPTS[role]}]
        for msg in messages:
            or_messages.append({
                "role": msg["role"],
                "content": msg["content"]
            })

        or_response = http_requests.post(
            "https://openrouter.ai/api/v1/chat/completions",
            headers={
                "Authorization": f"Bearer {OPENROUTER_API_KEY}",
                "Content-Type": "application/json"
            },
            json={
                "model": "openrouter/auto",
                "messages": or_messages,
                "max_tokens": 1000
            }
        )

        or_data = or_response.json()
        print(f"[OPENROUTER RESPONSE] {or_data}")

        if "error" in or_data:
            error_msg = or_data["error"].get("message", "Unknown error")
            return jsonify({"status": "error", "message": f"AI error: {error_msg}"}), 500

        reply = or_data["choices"][0]["message"]["content"]
        return jsonify({"status": "success", "reply": reply})

    except Exception as e:
        print(f"[CHATBOT ERROR] {str(e)}")
        return jsonify({"status": "error", "message": f"AI error: {str(e)}"}), 500

# ============================================
# RUN APP
# ============================================

if __name__ == '__main__':
    debug_mode = os.getenv("FLASK_DEBUG", "false").lower() == "true"
    host = os.getenv("FLASK_HOST", "127.0.0.1")
    port = int(os.getenv("FLASK_PORT", "5000"))
    app.run(debug=debug_mode, host=host, port=port)